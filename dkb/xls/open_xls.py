
from openpyxl import Workbook
from openpyxl import load_workbook

#from openpyxl import Workbook

class ExcelCfg(): # pylint: disable=missing-class-docstring, too-few-public-methods
    _gesamtkosten_table = {'start':'A1', 'stop': 'C1'}
    _account_wowa = {'start': 'A7', 'stop': 'C7'}
    xls_attr = {'as_template': False, 
                'macro': True,
                'extension' : 'xlsm'
                }

# pylint: disable=unused-argument, bare-except, missing-function-docstring
class ExcelWriter():
    '''
    class to write into existing Excel by using xlwings
    '''
    def __init__(self, xls_file) -> None:
        self.xls_file = xls_file
        # self.work_book = None
        # self.work_book = load_workbook(xls_file, read_only=False, keep_vba=True)
        self.work_book = Workbook()
        # -------------------------------------- #
 
    
    # def write_quick(self, sheet: str, data) -> bool:
    #     wb = Workbook(write_only=True)
    #     ws = wb.create_sheet(title=sheet)

    #     cells = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    #     for ln_ctr, rdata in enumerate(data, start=5):
    #         for cell, val in zip(cells, list(rdata)):
    #             sheet[cell+str(ln_ctr)] = val


    #     wb.save(self.xls_file)
    #     wb.close()
    #     return True
    
    def create_new_sheet(self, name: str, after: str):
        
        sheet = self.work_book.create_sheet(name)
        return sheet
    
    # def init_workbook(self):
    #     self.work_book = load_workbook(self.xls_file, read_only=False, keep_vba=True) 
    # FIXME: make more flexibil, otherwise not all data are written 
    def write_month(self, sheet, data):
        print(sheet)
        cells = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for ln_ctr, rdata in enumerate(data, start=5):
            for cell, val in zip(cells, list(rdata)):
                sheet[cell+str(ln_ctr)] = val
        self.work_book.save(self.xls_file)

class ExcelReader():
    """class to read data from created excel file and to provide data to db
    """
    def __init__(self, xls_file) -> None:
        self.xls_file = xls_file
        self.work_book = load_workbook(xls_file, read_only=True, keep_vba=True)
    
    # TODO: implement finder
    def _find_class_col(self, sheet) -> int:
        self.work_book.close()
        return 0

    def get_class_values(self, sheet):
        class_col = 0
        # class_col = self._find_class_col(sheet)
        start = 1
        ws = self.work_book[sheet] # pylint: disable=invalid-name
        values = ws.iter_rows(min_row=start, min_col=class_col, max_col=class_col, values_only=True)
        return values
        

    def quit_reader(self):
        self.work_book.close()