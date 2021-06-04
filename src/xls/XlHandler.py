#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
Created on 25.02.2016

@author: WowaN

https://openpyxl.readthedocs.io/en/stable/
https://openpyxl.readthedocs.io/en/stable/index.html
pip install openpyxl
'''
from abc import abstractproperty
import logging

from openpyxl import load_workbook

import xlrd as xl       # Excel read supports only xls


class XlRead():
    def __init__(self, sPath):
        self.book = load_workbook(sPath, read_only=False, keep_vba=False)
        self.activeSheet = None

    #=========================================================================
    # busyness functions
    #=========================================================================
    def save(self):
        sFileName = 'newSpec.xlsx'
        self.book.save(sFileName)
        logging.debug(
            f'XlHandler.save - Excel sheet saved: {sFileName}')

    def openSheet(self, sName):
        try:
            self.activeSheet = self.book[sName]
            return self.activeSheet
        except:
            #print('Excel sheet does not exist: ' + sName)
            logging.debug(
                f'XlHandler.openSheet - Excel sheet does not exist: {sName}')
            return None

    def getCellContant(self, iRow, iCol):
        #sCell = str(iCol)+str(iRow)
        if self.activeSheet != None:
            # return self.activeSheet[sCell].value
            oCell = self.activeSheet.cell(row=iRow, column=iCol)
            if self.activeSheet.cell(row=iRow, column=iCol).value != None:
                return self.activeSheet.cell(row=iRow, column=iCol).value
            else:
                return ''
        else:
            logging.debug('XlHandler.getCellContant - no active sheet')
            return None

    def searchRow(self, sValue, iStart):

        #rows = self.activeSheet.iter_rows(min_row=iStart, values_only=True)
        # for row in rows:
        for row in self.activeSheet.iter_rows(min_row=iStart):
            cl = row[0]
            if cl.value == sValue:
                logging.debug(
                    f'XlHandler.searchRow - found matching ID{cl.value}')
                #print('found matching ID', cl.value)
                return row
        return None

    def writeCell(self, row, iCol, sValue):
        logging.debug(f'XlHandler.writeCell - found matching ID: {row[iCol]}')
        row[iCol].value = sValue

    #=========================================================================
    # properties
    #=========================================================================

    @property
    def lastRow(self):
        if self.activeSheet:
            return self.activeSheet.max_row
        else:
            #print('sheet not opened')
            logging.debug('XlHandler.lastRow - sheet not opened')

    @property
    def lastColumn(self):
        if self.activeSheet:
            return self.activeSheet.max_column
        else:
            #print('sheet not opened')
            logging.debug('XlHandler.lastColumn - sheet not opened')
    #=========================================================================
    # about the class
    #=========================================================================

    def __repr__(self):
        return f"XlRead('{self.book}', '{self.activeSheet}')"

    def __str__(self):
        return f"XlRead('{self.book}', '{self.activeSheet}')"
